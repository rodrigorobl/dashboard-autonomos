import streamlit as st
import plotly.graph_objects as go
import openpyxl
import os

FILEPATH = 'dados/Planilha_de_Controle_de_Gastos_-_Autnomos.xlsx'

# ── Cores fixas (não mudam com o tema) ───────────────────────────────────────
C_GREEN  = '#16a34a'
C_RED    = '#dc2626'
C_AMBER  = '#d97706'
C_CYAN   = '#0284c7'
C_BLUE   = '#1e3a8a'
C_BLUE_L = '#3b82f6'

# ── Paletas light / dark ──────────────────────────────────────────────────────
LIGHT = dict(
    BG='#f8fafc', CARD='#ffffff', BORDER='#e2e8f0',
    TEXT='#1e293b', MUTED='#64748b',
    CHART_BG='#ffffff', CHART_GRID='#f1f5f9', CHART_LINE='#e2e8f0',
    TAB_BG='#ffffff', TAB_ACTIVE='#1e3a8a',
    TAB_TEXT='#64748b', BADGE_BG='#1e3a8a', BADGE_TEXT='#ffffff',
    HEADER_SUB='#94a3b8', DIVIDER='#e2e8f0',
    SCROLL_TRACK='#f8fafc', SCROLL_THUMB='#cbd5e1',
    SHADOW='rgba(0,0,0,0.06)',
)

DARK = dict(
    BG='#0f172a', CARD='#1e293b', BORDER='#334155',
    TEXT='#f1f5f9', MUTED='#94a3b8',
    CHART_BG='#1e293b', CHART_GRID='#263347', CHART_LINE='#334155',
    TAB_BG='#1e293b', TAB_ACTIVE='#3b82f6',
    TAB_TEXT='#94a3b8', BADGE_BG='#3b82f6', BADGE_TEXT='#ffffff',
    HEADER_SUB='#64748b', DIVIDER='#334155',
    SCROLL_TRACK='#0f172a', SCROLL_THUMB='#334155',
    SHADOW='rgba(0,0,0,0.25)',
)


def get_theme(dark: bool) -> dict:
    return DARK if dark else LIGHT


def get_chart_layout(dark: bool) -> dict:
    t = get_theme(dark)
    return dict(
        paper_bgcolor=t['CHART_BG'],
        plot_bgcolor=t['CHART_BG'],
        font=dict(color=t['TEXT'], family='Inter, sans-serif', size=12),
        title_font=dict(color=t['TEXT'], size=13, family='Inter, sans-serif'),
        xaxis=dict(gridcolor=t['CHART_GRID'], linecolor=t['CHART_LINE'],
                   tickfont=dict(color=t['MUTED'], size=10), zeroline=False),
        yaxis=dict(gridcolor=t['CHART_GRID'], linecolor=t['CHART_LINE'],
                   tickfont=dict(color=t['MUTED'], size=10), zeroline=False),
        legend=dict(bgcolor='rgba(0,0,0,0)', font=dict(color=t['MUTED'])),
        margin=dict(t=48, b=24, l=24, r=24),
    )


def apply_theme(dark: bool):
    t = get_theme(dark)
    st.markdown(f"""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');

    html, body, [class*="css"], .stApp, .main, .main > div {{
        background-color: {t['BG']} !important;
        font-family: 'Inter', sans-serif !important;
    }}

    .block-container {{
        padding-top: 2rem !important;
        max-width: 1280px !important;
    }}

    /* ── Sidebar ── */
    [data-testid="stSidebar"] {{
        background-color: {t['CARD']} !important;
        border-right: 1px solid {t['BORDER']} !important;
    }}
    [data-testid="stSidebar"] label,
    [data-testid="stSidebar"] p,
    [data-testid="stSidebar"] span {{
        color: {t['TEXT']} !important;
    }}

    /* ── Tabs ── */
    div[data-baseweb="tab-list"] {{
        background: {t['TAB_BG']} !important;
        border-radius: 8px !important;
        padding: 3px !important;
        border: 1px solid {t['BORDER']} !important;
        gap: 2px !important;
        box-shadow: 0 1px 3px {t['SHADOW']} !important;
    }}
    button[data-baseweb="tab"] {{
        font-family: 'Inter', sans-serif !important;
        font-size: 0.82rem !important;
        font-weight: 500 !important;
        color: {t['TAB_TEXT']} !important;
        background: transparent !important;
        border-radius: 6px !important;
        padding: 7px 24px !important;
        border: none !important;
    }}
    button[data-baseweb="tab"][aria-selected="true"] {{
        background: {t['TAB_ACTIVE']} !important;
        color: #ffffff !important;
    }}
    div[data-baseweb="tab-highlight"] {{ display: none !important; }}
    div[data-baseweb="tab-border"]    {{ display: none !important; }}

    /* ── Divider ── */
    hr {{ border-color: {t['DIVIDER']} !important; margin: 1.4rem 0 !important; }}

    /* ── Scrollbar ── */
    ::-webkit-scrollbar {{ width: 5px; }}
    ::-webkit-scrollbar-track {{ background: {t['SCROLL_TRACK']}; }}
    ::-webkit-scrollbar-thumb {{ background: {t['SCROLL_THUMB']}; border-radius: 4px; }}
    </style>
    """, unsafe_allow_html=True)


def kpi_card(label: str, value: str, accent: str, dark: bool = False) -> str:
    t = get_theme(dark)
    return f"""
    <div style="
        background: {t['CARD']};
        border: 1px solid {t['BORDER']};
        border-left: 4px solid {accent};
        border-radius: 8px;
        padding: 1.1rem 1.4rem;
        box-shadow: 0 1px 4px {t['SHADOW']};
        font-family: 'Inter', sans-serif;
    ">
      <div style="
        font-size: 0.68rem;
        font-weight: 600;
        color: {t['MUTED']};
        text-transform: uppercase;
        letter-spacing: 0.08em;
        margin-bottom: 0.55rem;
      ">{label}</div>
      <div style="
        font-size: 1.6rem;
        font-weight: 700;
        color: {t['TEXT']};
        letter-spacing: -0.02em;
        line-height: 1;
      ">{value}</div>
    </div>
    """


MONTH_NAMES = ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun',
               'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']

CATEGORY_HEADERS = {
    'Alimentação', 'Moradia', 'Educação', 'Animal de Estimação',
    'Saúde', 'Transporte', 'Pessoal', 'Lazer', 'Serviços Financeiros'
}


def format_month(dt):
    """Converte datetime para string 'Mmm/AA'. Retorna str(dt) se não for datetime."""
    if hasattr(dt, 'month'):
        return f"{MONTH_NAMES[dt.month - 1]}/{str(dt.year)[2:]}"
    return str(dt)


def load_pj_data(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['Pessoa Jurídica']
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    months = [format_month(r) for r in rows[0][1:13]]

    receitas = {}
    despesas = {}
    totais_receita = []
    totais_despesa = []
    resultado = []
    current_section = None

    for row in rows[1:]:
        if row[0] is None:
            continue
        name = str(row[0])
        values = [v if v is not None else 0 for v in row[1:len(months)+1]]

        if name == 'Receitas':
            current_section = 'receitas'
        elif name == 'Despesas':
            current_section = 'despesas'
        elif name == 'Total de Receitas':
            totais_receita = values
        elif name == 'Total das Despesas':
            totais_despesa = values
        elif name.startswith('Resultado Operacional'):
            resultado = values
        elif name.startswith('Alterar'):
            continue
        elif name.startswith('- '):
            item_name = name[2:]
            if current_section == 'receitas':
                receitas[item_name] = values
            elif current_section == 'despesas':
                despesas[item_name] = values

    return {
        'months': months,
        'receitas': receitas,
        'despesas': despesas,
        'totais_receita': totais_receita,
        'totais_despesa': totais_despesa,
        'resultado': resultado,
    }


def load_pf_data(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['Pessoa Física']
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    val_indices = list(range(1, 25, 2))
    pay_indices = list(range(2, 26, 2))

    months = [format_month(rows[0][i]) for i in val_indices]

    categories = {}
    payment_totals = {}
    totais_despesa = []
    renda = []
    resultado = []
    investimentos = []
    current_category = None

    def get_vals(row):
        return [row[i] if i < len(row) and row[i] is not None else 0 for i in val_indices]

    def get_methods(row):
        return [row[i] if i < len(row) and row[i] is not None else '' for i in pay_indices]

    for row in rows[1:]:
        if row[0] is None:
            continue
        name = str(row[0])

        if name in ('Despedas', 'Despesas'):
            continue
        if name == 'Alterar somente os campos em azul':
            continue

        if name in CATEGORY_HEADERS:
            current_category = name
            categories[current_category] = {}
            continue

        if name == 'Total das Despesas':
            totais_despesa = get_vals(row)
            continue
        if name == 'Renda Mensal':
            renda = get_vals(row)
            continue
        if name.startswith('Resultado Operacional'):
            resultado = get_vals(row)
            continue
        if name == 'Investimentos Mensais':
            investimentos = get_vals(row)
            continue

        if current_category is not None:
            vals = get_vals(row)
            methods = get_methods(row)
            categories[current_category][name] = {'values': vals, 'methods': methods}
            for v, m in zip(vals, methods):
                if m and v:
                    payment_totals[m] = payment_totals.get(m, 0) + v

    return {
        'months': months,
        'categories': categories,
        'payment_totals': payment_totals,
        'totais_despesa': totais_despesa,
        'renda': renda,
        'resultado': resultado,
        'investimentos': investimentos,
    }


def _apply_chart_layout(fig, title, dark: bool, **extra):
    layout = get_chart_layout(dark)
    fig.update_layout(title=title, **{**layout, **extra})
    return fig


def render_pj_tab(data, dark: bool):
    t = get_theme(dark)

    total_receita = sum(data['totais_receita'])
    total_despesa = sum(data['totais_despesa'])
    resultado_total = sum(data['resultado'])
    melhor_idx = data['resultado'].index(max(data['resultado']))
    melhor_mes = data['months'][melhor_idx]

    sinal = '+' if resultado_total >= 0 else ''
    cor_resultado = C_GREEN if resultado_total >= 0 else C_RED
    col1, col2, col3, col4 = st.columns(4)
    col1.markdown(kpi_card('Receita Total',
                            f"R$ {total_receita:,.0f}".replace(',', '.'), C_GREEN, dark),
                  unsafe_allow_html=True)
    col2.markdown(kpi_card('Despesas Total',
                            f"R$ {total_despesa:,.0f}".replace(',', '.'), C_RED, dark),
                  unsafe_allow_html=True)
    col3.markdown(kpi_card('Resultado Operacional',
                            f"{sinal}R$ {resultado_total:,.0f}".replace(',', '.'), cor_resultado, dark),
                  unsafe_allow_html=True)
    col4.markdown(kpi_card('Melhor Mês', melhor_mes, C_CYAN, dark),
                  unsafe_allow_html=True)

    st.divider()

    col_l, col_r = st.columns(2)

    with col_l:
        fig = go.Figure()
        fig.add_bar(x=data['months'], y=data['totais_receita'],
                    name='Receitas', marker_color=C_GREEN, marker_opacity=0.9)
        fig.add_bar(x=data['months'], y=data['totais_despesa'],
                    name='Despesas', marker_color=C_RED, marker_opacity=0.9)
        _apply_chart_layout(fig, 'Receitas vs Despesas por Mês', dark,
                            barmode='group',
                            legend=dict(orientation='h', yanchor='bottom',
                                        y=1.02, xanchor='right', x=1,
                                        bgcolor='rgba(0,0,0,0)',
                                        font=dict(color=t['MUTED'])))
        st.plotly_chart(fig, use_container_width=True)

    with col_r:
        desp_names = list(data['despesas'].keys())
        desp_totals = [sum(data['despesas'][k]) for k in desp_names]
        pairs = [(n, v) for n, v in zip(desp_names, desp_totals) if v > 0]
        if pairs:
            names, vals = zip(*pairs)
            pie_line_color = t['CARD']
            fig = go.Figure(go.Pie(
                labels=list(names), values=list(vals),
                hole=0.45,
                textposition='inside', textinfo='percent',
                marker=dict(colors=[
                    '#1e3a8a','#3b82f6','#0284c7','#0ea5e9',
                    '#16a34a','#22c55e','#d97706','#dc2626','#7c3aed','#64748b'
                ], line=dict(color=pie_line_color, width=2))
            ))
            fig.update_layout(
                title='Composição das Despesas',
                **{**get_chart_layout(dark),
                   'legend': dict(bgcolor='rgba(0,0,0,0)', font=dict(color=t['MUTED'], size=10))}
            )
            st.plotly_chart(fig, use_container_width=True)

    col_l2, col_r2 = st.columns(2)

    with col_l2:
        cores_res = [C_GREEN if v >= 0 else C_RED for v in data['resultado']]
        fig = go.Figure()
        fig.add_trace(go.Scatter(
            x=data['months'], y=data['resultado'],
            mode='lines+markers',
            line=dict(color=C_BLUE_L if dark else C_BLUE, width=2.5),
            marker=dict(color=cores_res, size=9, line=dict(color=t['CARD'], width=1.5)),
            fill='tozeroy',
            fillcolor='rgba(59,130,246,0.08)' if dark else 'rgba(30,58,138,0.06)'
        ))
        fig.add_hline(y=0, line_dash='dot', line_color=t['BORDER'], opacity=0.8)
        _apply_chart_layout(fig, 'Resultado Operacional por Mês', dark,
                            yaxis=dict(**get_chart_layout(dark)['yaxis'], title='R$'))
        st.plotly_chart(fig, use_container_width=True)

    with col_r2:
        desp_sorted = sorted(
            [(n, sum(data['despesas'][n])) for n in data['despesas']
             if sum(data['despesas'][n]) > 0],
            key=lambda x: x[1], reverse=True
        )
        if desp_sorted:
            names, vals = zip(*desp_sorted)
            n = len(vals)
            if dark:
                bar_colors = [f'rgba(59,130,246,{0.4 + 0.6*(i/max(n-1,1)):.2f})' for i in range(n)]
            else:
                bar_colors = [f'rgba(30,58,138,{0.4 + 0.6*(i/max(n-1,1)):.2f})' for i in range(n)]
            fig = go.Figure(go.Bar(
                x=list(vals), y=list(names),
                orientation='h',
                marker=dict(color=bar_colors),
                text=[f'R$ {v:,.0f}'.replace(',', '.') for v in vals],
                textposition='outside',
                textfont=dict(color=t['MUTED'], size=10, family='IBM Plex Mono, monospace')
            ))
            _apply_chart_layout(fig, 'Top Categorias de Despesa', dark,
                                xaxis=dict(**get_chart_layout(dark)['xaxis'], title='Total (R$)'),
                                yaxis=dict(**get_chart_layout(dark)['yaxis'], autorange='reversed'))
            st.plotly_chart(fig, use_container_width=True)


def render_pf_tab(data, dark: bool):
    t = get_theme(dark)

    total_renda = sum(data['renda'])
    total_despesa = sum(data['totais_despesa'])
    total_invest = sum(data['investimentos'])
    meio_mais_usado = (
        max(data['payment_totals'], key=data['payment_totals'].get)
        if data['payment_totals'] else 'N/A'
    )

    col1, col2, col3, col4 = st.columns(4)
    col1.markdown(kpi_card('Renda Total',
                            f"R$ {total_renda:,.0f}".replace(',', '.'), C_GREEN, dark),
                  unsafe_allow_html=True)
    col2.markdown(kpi_card('Despesas Total',
                            f"R$ {total_despesa:,.0f}".replace(',', '.'), C_RED, dark),
                  unsafe_allow_html=True)
    col3.markdown(kpi_card('Investimentos',
                            f"R$ {total_invest:,.0f}".replace(',', '.'), C_AMBER, dark),
                  unsafe_allow_html=True)
    col4.markdown(kpi_card('Meio + Usado', meio_mais_usado, C_CYAN, dark),
                  unsafe_allow_html=True)

    st.divider()

    col_l, col_r = st.columns(2)

    with col_l:
        fig = go.Figure()
        fig.add_bar(x=data['months'], y=data['renda'],
                    name='Renda', marker_color=C_GREEN, marker_opacity=0.9)
        fig.add_bar(x=data['months'], y=data['totais_despesa'],
                    name='Despesas', marker_color=C_RED, marker_opacity=0.9)
        _apply_chart_layout(fig, 'Renda vs Despesas por Mês', dark,
                            barmode='group',
                            legend=dict(orientation='h', yanchor='bottom',
                                        y=1.02, xanchor='right', x=1,
                                        bgcolor='rgba(0,0,0,0)',
                                        font=dict(color=t['MUTED'])))
        st.plotly_chart(fig, use_container_width=True)

    with col_r:
        cat_totals = {}
        for cat_name, items in data['categories'].items():
            total = sum(sum(item['values']) for item in items.values())
            if total > 0:
                cat_totals[cat_name] = total
        if cat_totals:
            pie_line_color = t['CARD']
            fig = go.Figure(go.Pie(
                labels=list(cat_totals.keys()),
                values=list(cat_totals.values()),
                hole=0.45,
                textposition='inside', textinfo='percent',
                marker=dict(colors=[
                    '#00d4ff','#10b981','#f59e0b','#ef4444',
                    '#8b5cf6','#ec4899','#14b8a6','#f97316','#06b6d4'
                ], line=dict(color=pie_line_color, width=2))
            ))
            fig.update_layout(
                title='Despesas por Categoria',
                **{**get_chart_layout(dark),
                   'legend': dict(bgcolor='rgba(0,0,0,0)', font=dict(color=t['MUTED'], size=10))}
            )
            st.plotly_chart(fig, use_container_width=True)

    col_l2, col_r2 = st.columns(2)

    with col_l2:
        if data['payment_totals']:
            pay_sorted = sorted(data['payment_totals'].items(), key=lambda x: x[1], reverse=True)
            methods, vals = zip(*pay_sorted)
            n = len(vals)
            bar_colors = [f'rgba(2,132,199,{0.4 + 0.6*(i/max(n-1,1)):.2f})' for i in range(n)]
            fig = go.Figure(go.Bar(
                x=list(vals), y=list(methods),
                orientation='h',
                marker=dict(color=bar_colors),
                text=[f'R$ {v:,.0f}'.replace(',', '.') for v in vals],
                textposition='outside',
                textfont=dict(color=t['MUTED'], size=10, family='IBM Plex Mono, monospace')
            ))
            _apply_chart_layout(fig, 'Gastos por Meio de Pagamento', dark,
                                xaxis=dict(**get_chart_layout(dark)['xaxis'], title='Total (R$)'),
                                yaxis=dict(**get_chart_layout(dark)['yaxis'], autorange='reversed'))
            st.plotly_chart(fig, use_container_width=True)

    with col_r2:
        saldo = [r - d for r, d in zip(data['renda'], data['totais_despesa'])]
        cores_saldo = [C_GREEN if s >= 0 else C_RED for s in saldo]
        fig = go.Figure(go.Bar(
            x=data['months'], y=saldo,
            marker_color=cores_saldo,
            text=[f"R$ {s:,.0f}".replace(',', '.') for s in saldo],
            textposition='outside',
            textfont=dict(color=t['MUTED'], size=10, family='IBM Plex Mono, monospace')
        ))
        fig.add_hline(y=0, line_dash='dot', line_color=t['BORDER'], opacity=0.8)
        _apply_chart_layout(fig, 'Saldo Mensal (Renda − Despesas)', dark,
                            yaxis=dict(**get_chart_layout(dark)['yaxis'], title='R$'))
        st.plotly_chart(fig, use_container_width=True)


def main():
    st.set_page_config(
        page_title='Dashboard Autônomos',
        page_icon='📊',
        layout='wide',
        initial_sidebar_state='collapsed'
    )

    # ── Toggle tema ───────────────────────────────────────────────────────────
    dark = st.sidebar.toggle('🌙 Modo escuro', value=False)

    apply_theme(dark)

    t = get_theme(dark)
    st.markdown(f"""
    <div style="
        background:{t['CARD']};border:1px solid {t['BORDER']};border-radius:8px;
        padding:1.2rem 1.6rem;margin-bottom:1.2rem;
        box-shadow:0 1px 4px {t['SHADOW']};
        display:flex;align-items:center;justify-content:space-between;
    ">
      <div>
        <div style="
            font-family:'Inter',sans-serif;font-weight:700;font-size:1.4rem;
            color:{t['TEXT']};letter-spacing:-0.01em;margin-bottom:2px;
        ">Controle de Gastos — Autônomos</div>
        <div style="
            font-family:'Inter',sans-serif;font-size:0.75rem;
            color:{t['HEADER_SUB']};font-weight:400;
        ">Planilha de Controle de Gastos · 2018</div>
      </div>
      <div style="
          background:{t['BADGE_BG']};color:{t['BADGE_TEXT']};
          font-family:'Inter',sans-serif;font-size:0.72rem;font-weight:600;
          padding:5px 14px;border-radius:20px;letter-spacing:0.02em;
      ">FINANCEIRO</div>
    </div>
    """, unsafe_allow_html=True)

    # ── Fonte de dados ────────────────────────────────────────────────────────
    uploaded = st.sidebar.file_uploader(
        'Atualizar planilha',
        type=['xlsx'],
        help='Faça upload de uma nova versão da planilha para atualizar o dashboard.'
    )

    if uploaded is not None:
        source = uploaded
    elif os.path.exists(FILEPATH):
        source = FILEPATH
    else:
        st.sidebar.warning('Nenhuma planilha encontrada.')
        st.info('Use o painel lateral para fazer upload da planilha (.xlsx).')
        st.stop()

    pj_data = load_pj_data(source)
    pf_data = load_pf_data(source)

    tab_pj, tab_pf = st.tabs(['📊 Pessoa Jurídica', '🏠 Pessoa Física'])
    with tab_pj:
        render_pj_tab(pj_data, dark)
    with tab_pf:
        render_pf_tab(pf_data, dark)


if __name__ == '__main__':
    main()
